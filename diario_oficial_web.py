from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Função para salvar dados no Excel usando Pandas
def salvar_dados_excel_pandas(resultados, nome_arquivo):
    try:
        # Salvar o DataFrame em um arquivo Excel
        resultados.to_excel(nome_arquivo, index=False)
        print(f"Os resultados foram salvos no arquivo '{nome_arquivo}'.")
    except Exception as e:
        print(f"Erro ao salvar os resultados no arquivo '{nome_arquivo}': {e}")

# Configurações iniciais
clientes = pd.read_excel('clientes_caceal.xlsx')
service = Service()
driver = webdriver.Chrome(service=service)

periodo_inicial = input("Período Inicial: ")
periodo_final = input("Período Final: ")

print(f'Periodo {periodo_inicial} e {periodo_final}')

# Abre uma página da web
driver.get("https://diario.imprensaoficial.al.gov.br")
driver.maximize_window()
pbar = tqdm(total=len(clientes['caceal']), position=0, leave=True)
localizado = 0
nao_localizado = 0

# Lista para armazenar as linhas de resultado
resultados_lista = []

# Função para salvar dados no Excel usando openpyxl
def salvar_dados_excel_openpyxl(razao_social, codigo_caceal, info_diario_texto, file_path):
    try:
        # Tenta carregar a planilha existente
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        # Se o arquivo não existir, cria uma nova planilha
        workbook = Workbook()

    # Seleciona a planilha ativa (por padrão, a primeira planilha)
    sheet = workbook.active

    # Adiciona os cabeçalhos se a planilha estiver vazia
    if sheet.max_row == 0:
        sheet.append(['Razao_social', 'codigo_caceal', 'info_diario_texto'])

    # Adiciona os dados na próxima linha disponível
    sheet.append([razao_social, codigo_caceal, info_diario_texto])

    # Salva as alterações no arquivo
    workbook.save(file_path)

# Iteração sobre clientes
for cliente in clientes.index:
    pbar.update()
    codigo_caceal = clientes.loc[cliente, "caceal"]
    nome_cliente = clientes.loc[cliente, "Razao_social"]
    
    # Realizar a busca
    procurar = driver.find_element(By.XPATH, '//*[@id="inputPalavraChave"]').send_keys(codigo_caceal)
    time.sleep(3)
    data_incial = driver.find_element(By.ID, 'inputPeriodoInicial').send_keys(periodo_inicial)
    time.sleep(1)
    data_final = driver.find_element(By.ID, 'inputPeriodoFinal').send_keys(periodo_final)
    time.sleep(1)

    busca = driver.find_element(By.CSS_SELECTOR, 'button.btn.btn-primary')
    busca.click()

    time.sleep(3)
    resultados_elementos = driver.find_elements(By.XPATH, '//*[@id="corpo-site"]/div/div[6]/div[2]/div[1]/div/div[1]/div/a[2]/span')
    info_diario = driver.find_elements(By.TAG_NAME, 'h6')
    razao_social = driver.find_elements(By.CLASS_NAME, 'highlight')

    if resultados_elementos:
        localizado += 1

        resultado_texto = resultados_elementos[0].text
        info_diario_texto = info_diario[0].text
        if len(razao_social)>1:
            razao_social_texto = razao_social[1].text
            print(f'_Encontrado {nome_cliente} no {codigo_caceal}\n{info_diario_texto}')
        else:
            razao_social= "N/A"
            print(f"Atenção: Não foi possível encontrar a razão social para o cliente {nome_cliente} com CACEAL {codigo_caceal}")
        # Adicionar resultados à lista
        nova_linha = {"Razão social": nome_cliente, "Caceal": codigo_caceal, "Informacao Diario": info_diario_texto}
        resultados_lista.append(nova_linha)
        
        # Salvar dados usando openpyxl
        data_salvamento = datetime.now().strftime('%d-%m-%Y')
        salvar_dados_excel_openpyxl(nome_cliente, codigo_caceal, info_diario_texto, f'dados_encontrados_{data_salvamento}.xlsx')
        
        procurar = driver.find_element(By.XPATH, '//*[@id="inputPalavraChave"]').clear()
        time.sleep(1)
    else:
        nao_localizado += 1
        print(f'_Empresas não encontradas {nao_localizado}')
        print(f"_Não foi encontrada a empresa {nome_cliente} {codigo_caceal}")
        time.sleep(1)
        procurar = driver.find_element(By.XPATH, '//*[@id="inputPalavraChave"]').clear()
        print('Buscando nova empresa')

# Transformar a lista de resultados em DataFrame
resultados = pd.DataFrame(resultados_lista)

# Salvar todos os resultados usando Pandas
nome_arquivo = f"dados_encontrados_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
salvar_dados_excel_pandas(resultados, nome_arquivo)
